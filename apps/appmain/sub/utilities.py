import io
import re
from datetime import datetime as dt
from pathlib import Path
from time import time

import numpy as np
import openpyxl
import pandas as pd

mid = Path(__file__).name


# # decorator test
# def log_func(func):
#     def inner(*args):
#         print(f"----- decorator test Start: {dt.fromtimestamp(time())} -----")
#         print(f"{func.__name__= }")
#         print(f"{args= }")
#         result = func(*args)
#         print(f"----- decorator test End:   {dt.fromtimestamp(time())} -----")
#         # return func(*args)
#         return result
#     return inner
#
#
# @log_func
def validity_check_control_data_dict(control_data_dict):
    print(f"[{mid}] validity_check_control_data_dict")

    # check key_type
    key_type = control_data_dict["key_type"]
    allowed_words = ["text", "integer", "real"]
    allowed_words += ["varchar", "timestamp"]  # PostgreSQL ex.) varchar(100) -> varchar

    # if False in [True if item in allowed_words else False for item in key_type]:
    if False in [True if re.sub('\([0-9]*\)', '', item) in allowed_words else False for item in key_type]:
        alert = ('NG', 'error: valid key_type: text, integer, real, varchar, timestamp')
        return alert

    # check key
    key = control_data_dict["key"]
    not_allowed_words = ["table", "replace", "into", "insert", "values", "from", "order", "by", "limit"]
    not_allowed_words += ["update_at", "modified_at"]

    for word in not_allowed_words:
        if word in key:
            alert = ('NG', f'error: "key" contains invalid name: {word}')
            return alert

    alert = ("OK", "Done.")
    return alert


def gen_control_data_dict(data, file_name):
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, keep_vba=True, data_only=True)
    try:
        control_sheet = wb["control"]
    except Exception as err:
        alert = ("NG", f"error: control sheet read error. {err= }")
        return alert, {}  # alert, control_data_dict

    worksheet_name = []
    cell_position = []
    key = []
    key_type = []
    pk_mark = []

    ptn1 = re.compile('\(.+?\)')  # 括弧内抽出
    ptn2 = re.compile('[ /.%　]')  # slash period percent 全角スペース

    for row in control_sheet.iter_rows(min_row=2):
        worksheet_name.append(row[0].value)
        cell_position.append(row[1].value)
        s = row[2].value
        s = re.sub(ptn1, "", s)
        s = re.sub(ptn2, "", s)
        key.append(s)
        key_type.append(row[3].value)
        pk_mark.append(row[4].value)

    control_data_dict = dict()
    control_data_dict["file_name"] = file_name[:-5]  # remove ".xlsx", file_name means collection or table
    control_data_dict["worksheet_name"] = worksheet_name
    control_data_dict["cell_position"] = cell_position
    control_data_dict["key"] = key
    control_data_dict["key_type"] = key_type

    # set primary key
    primary_key_info = []
    for i, k in enumerate(pk_mark):
        if k:
            primary_key_info.append((k, i, key[i]))
        primary_key_info.sort(key=lambda x: x[0])
    control_data_dict["primary_key_info"] = primary_key_info  # ex.) ((1,3,'company'), (2,4,'section'))

    primary_key = [item[2] for item in primary_key_info]
    control_data_dict["primary_key"] = primary_key

    # set form type
    form_type_check = [x for x in cell_position if re.findall("[+]", x)]  # "+" exists?
    form_type = "2" if form_type_check else "1"  # "2" 明細タイプ "1" 帳票タイプ
    control_data_dict["form_type"] = form_type

    alert = validity_check_control_data_dict(control_data_dict)

    return alert, control_data_dict


def gen_control_data_dict_for_csv(data, file_name):
    df = pd.read_csv(io.BytesIO(data), encoding="utf-8", header=0)
    col_dtype = pd.DataFrame(zip(df.columns, df.dtypes.values.tolist()), columns=["col", "dtype"])
    col_dtype["dtype"].replace(["int64", "float64", "object"], ["integer", "real", "text"], inplace=True)

    control_data_dict = dict()
    control_data_dict["file_name"] = file_name[:-4]  # remove ".csv", file_name means collection or table
    control_data_dict["worksheet_name"] = []
    control_data_dict["cell_position"] = []
    control_data_dict["key"] = col_dtype["col"].values.tolist()
    control_data_dict["key_type"] = col_dtype["dtype"].values.tolist()
    control_data_dict["primary_key_info"] = []
    control_data_dict["primary_key"] = []
    control_data_dict["form_type"] = "3"  # .csv file type

    alert = ("OK", "Done.")
    return alert, control_data_dict


# cell-oriented Excel format
def gen_input_data_type1(form_sheet, control_data_dict):
    print(f"[{mid}] gen_input_data_type1")

    cell_position = control_data_dict["cell_position"]

    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    # setup cell position, ex. "AF40" -> (39, 31)
    cell_position_dim2 = []

    for cell in cell_position:
        cell = cell.upper()
        idx = cell.index(re.findall("[0-9]", cell)[0])  # "AF40" -> "AF", "40"

        row = int(cell[idx:]) - 1  # 39 = 40 - 1

        x = cell[:idx]
        c = x[-1]
        col = (len(x) - 1) * len(alphabet) + alphabet.index(c)  # "AF" -> 26 + 5 = 31

        cell_position_dim2.append((row, col))

    # setup input data
    key = control_data_dict["key"]
    key_type = control_data_dict["key_type"]

    kv = {}

    for i, cell in enumerate(cell_position_dim2):
        v = form_sheet.iat[cell]

        if "text" in key_type[i]:
            kv[key[i]] = str(v).replace(' ', '')
        elif "integer" in key_type[i]:
            kv[key[i]] = int(v)
        else:  # real
            # print(f"**** {type(v)= } {v= }")
            # kv[key[i]] = v if v == v else 0.0
            kv[key[i]] = v

    # check the first primary key is not null
    primary_key_info = control_data_dict["primary_key_info"]
    if [bool(pk[2]) if kv[pk[2]] != "nan" else False for pk in primary_key_info].count(False) != 0:
        kv = {}  # if primary key is null, then reset kv

    return kv


# record-oriented Excel format
def gen_input_data_type2(form_sheet, control_data_dict):
    cell_position = control_data_dict["cell_position"]
    primary_key_info = control_data_dict["primary_key_info"]

    # calculate begin_row_index, max_column_index and cols
    cp = [x[:-1] for x in cell_position]  # remove "+"

    cp_sorted = sorted(sorted(cp), key=len)  # alphabetical and length sort
    begin_row_index = int(re.split('[a-zA-Z]', cp_sorted[0])[1])  # ex.) "A2" -> 2

    max_column_name = re.split('[0-9]', cp_sorted[-1])[0]  # ex.) "AE2" -> "AE"

    alphabet = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")  # A - Z
    c1 = [c1 + c2 for c1 in alphabet for c2 in alphabet]  # AA - ZZ
    c2 = [c1 + c2 + c3 for c1 in alphabet for c2 in alphabet for c3 in alphabet]  # AAA - ZZZ
    c = alphabet + c1 + c2  # A - ZZZ

    max_column_index = c.index(max_column_name) + 1

    cols = [re.split('[0-9]', x)[0] for x in cp]

    # create mask
    mask = [True if s in cols else False for s in c[:max_column_index]]

    rows = list()

    max_records = 500_000

    # iter_rows is faster than "in form_sheet"
    for record in form_sheet.iter_rows(min_row=begin_row_index, max_col=max_column_index, max_row=max_records):
        items = [item.value for item in record]
        items = np.array(items)
        value = items[mask]

        # check the primary key(s) is(are) not null.
        primary_key_null = False
        for pk in primary_key_info:
            if not value[pk[1]]:  # ex.) primary_key_info = ((1, 3, 'company'), (2, 4, 'section'))
                primary_key_null = True
                break
        if primary_key_null:
            break

        rows.append(value.tolist())

    return rows


def gen_input_data(data, control_data_dict):
    form_type = control_data_dict["form_type"]
    alert = ("OK", "Done.")  # initialization
    try:
        if form_type == "1":  # cell-oriented Excel format
            worksheet_name = list(set(control_data_dict["worksheet_name"]))[0]
            form_sheet = pd.read_excel(io.BytesIO(data), sheet_name=worksheet_name, header=None)
            input_data = gen_input_data_type1(form_sheet, control_data_dict)
            input_data = [list(input_data.values())]  # list in list
        elif form_type == "2":  # record-oriented Excel format
            worksheet_name = list(set(control_data_dict["worksheet_name"]))[0]
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, keep_vba=True, data_only=True)
            form_sheet = wb[worksheet_name]
            input_data = gen_input_data_type2(form_sheet, control_data_dict)
        elif form_type == "3":  # csv format
            df = pd.read_csv(io.BytesIO(data), encoding="utf-8", header=1)
            input_data = df.values.tolist()
        else:
            input_data = ""
    except Exception as err:
        input_data = ""
        alert = ("NG", f"error: gen_input_data error occurred. {err= }")

    return alert, input_data
